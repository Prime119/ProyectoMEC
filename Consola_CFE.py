import sys
import time
import numpy as np
import pandas as pd
import subprocess
import serial
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QGridLayout, QSplitter, QPushButton, QTableWidget,
    QTableWidgetItem, QCheckBox, QDoubleSpinBox, QGroupBox,
    QScrollArea, QSlider, QComboBox, QFrame, QSizePolicy, QHeaderView,
    QAbstractItemView, QTextEdit
)
from PyQt6.QtCore import QTimer, Qt, QSize, QFileSystemWatcher
from PyQt6.QtGui import QFont, QColor, QPalette, QLinearGradient, QPainter, QBrush, QPen
import pyqtgraph as pg
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
import os

# TENSORFLOW: Silenciar logs antes de importar
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'

import tensorflow as tf
import keras
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import MinMaxScaler
import warnings
warnings.filterwarnings('ignore')

pg.setConfigOption('background', '#070B14')
pg.setConfigOption('foreground', '#9CA3AF')
pg.setConfigOptions(antialias=True)

C_BG      = '#070B14'
C_PANEL   = '#0C1320'
C_BORDER  = '#16304a'
C_TEXT_M  = '#9CA3AF'
C_TEXT_H  = '#F3F4F6'
C_MAIN    = '#22D3EE'
C_RMS_BLU = '#1E40AF'
C_ACCENT  = '#FACC15'
C_ORANGE  = '#F97316'
C_PURPLE  = '#A855F7'
C_WHITE   = '#FFFFFF'
C_NEUTRAL = '#6B7280'
C_GOOD    = '#10B981'
C_CRIT    = '#EF4444'

# --- Paleta HUD futurista (estilo J.A.R.V.I.S. / Transformer) ---
C_HUD     = '#22D3EE'   # cian principal
C_HUD_DK  = '#0E7490'   # cian oscuro (bordes tenues)
C_GLOW    = '#67E8F9'   # cian brillante (acentos/realce)
C_HUD_BG  = '#091824'   # fondo de paneles HUD


# --- Helpers de estilo holográfico para las gráficas (glow + áreas translúcidas) ---
def _holo_pen(color, width=2.0):
    return pg.mkPen(color, width=width)

def _glow_pen(color, width=8, alpha=70):
    c = pg.mkColor(color); c.setAlpha(alpha)
    return pg.mkPen(c, width=width)

def _area_brush(color, alpha=40):
    c = pg.mkColor(color); c.setAlpha(alpha)
    return pg.mkBrush(c)

C_VOLT    = C_MAIN
C_CURR    = C_ACCENT
C_P_ACT   = C_MAIN
C_P_REA   = C_ACCENT
C_P_APP   = '#D1D5DB'
C_VIB     = C_PURPLE
C_FREQ    = C_NEUTRAL
C_RMS     = C_RMS_BLU
C_THD     = C_ORANGE
C_PF      = C_WHITE
C_SALUD   = C_GOOD
C_TEMP    = C_CRIT

STYLE_MAIN = f"""
QMainWindow, QWidget {{ background:{C_BG}; color:{C_TEXT_M}; font-family:'Segoe UI', sans-serif; }}
QLabel {{ color:{C_TEXT_M}; }}
QSlider::groove:horizontal {{ height:2px; background:{C_BORDER}; border-radius:1px; }}
QDoubleSpinBox {{ background:#1F2937; border:1px solid {C_BORDER}; color:{C_TEXT_H}; font-family:'Consolas'; padding:2px; border-radius:2px; }}
QPushButton {{ background:#1F2937; border:1px solid {C_BORDER}; color:{C_TEXT_M}; padding:4px 12px; border-radius:2px; font-weight:bold; }}
QPushButton:hover {{ background:#374151; color:{C_TEXT_H}; }}
QTableWidget {{ background:{C_PANEL}; alternate-background-color:#0a1a26; color:{C_TEXT_H}; gridline-color:{C_BORDER}; font-family:'Consolas', monospace; font-size:10px; border:none; }}
QTableWidget QHeaderView::section {{ background:#0a1c2b; color:{C_GLOW}; font-family:'Segoe UI'; font-size:9px; font-weight:bold; text-transform:uppercase; letter-spacing:1px; padding:4px; border:none; border-bottom:1px solid {C_HUD_DK}; border-right:1px solid {C_BORDER}; }}
QScrollBar:vertical {{ background:{C_BG}; width:8px; }}
QScrollBar::handle:vertical {{ background:{C_HUD_DK}; border-radius:4px; }}
QScrollBar::handle:vertical:hover {{ background:{C_HUD}; }}
QTextEdit {{ background:{C_HUD_BG}; color:#E5F6FF; border:1px solid {C_HUD}; selection-background-color:{C_HUD_DK}; font-family:'Consolas', monospace; font-size:11px; }}
"""

# MOTOR DE IA TENSORFLOW — AUTOENCODER + PREDICTOR LSTM
class TensorFlowBrain:
    """
    Motor de IA dual con TensorFlow:
    1. AUTOENCODER: Detecta anomalías comparando la reconstrucción de señales.
       Si la señal reconstruida difiere mucho de la real, hay una anomalía.
    2. LSTM PREDICTOR: Red neuronal recurrente que predice el siguiente valor
       de salud del sistema basado en el historial de los últimos 20 ticks.
    3. ISOLATION FOREST (sklearn): Capa adicional de detección de outliers.
    """

    # Número de features de entrada: V, I, P, Q, PF, THD, Vib, Temp
    N_FEATURES = 8
    # Ventana de tiempo para el LSTM (últimos N ticks)
    SEQ_LEN    = 20
    # Muestras mínimas para entrenar
    MIN_SAMPLES = 60

    def __init__(self):
        self.scaler        = MinMaxScaler()
        self.iso_forest    = IsolationForest(contamination=0.05, n_estimators=100, random_state=42)
        self.is_trained    = False
        self.training_buf  = []          # buffer crudo de features
        self.seq_buf       = []          # buffer de secuencias para LSTM

        # Historial de métricas TF para mostrar en gráfica
        self.hist_loss_ae  = []          # pérdida del autoencoder en cada tick
        self.hist_pred_sal = []          # salud predicha por LSTM
        self.hist_anomaly  = []          # score de anomalía normalizado

        # Umbrales adaptativos (se calibran durante entrenamiento)
        self.threshold_ae  = 0.05        # error de reconstrucción máximo normal
        self.mean_loss     = 0.0
        self.std_loss      = 1.0

        # Construir modelos
        self._build_autoencoder()
        self._build_lstm()

        # Estado del sistema
        self.estado        = "APRENDIENDO"
        self.confianza     = 0.0         # 0-100%
        self.pred_salud    = 100.0
        self.anomaly_score = 0.0
        self.ae_loss       = 0.0
        self.n_anomalias   = 0
        self.samples_seen  = 0

    def _build_autoencoder(self):
        """
        Autoencoder denso: comprime 8 features → 4 → 2 → 4 → 8
        Si la reconstrucción falla, la señal es anómala.
        """
        inp = keras.Input(shape=(self.N_FEATURES,), name="ae_input")
        x   = keras.layers.Dense(16, activation='relu', name="enc1")(inp)
        x   = keras.layers.Dense(8,  activation='relu', name="enc2")(x)
        x   = keras.layers.Dense(4,  activation='relu', name="bottleneck")(x)
        x   = keras.layers.Dense(8,  activation='relu', name="dec1")(x)
        out = keras.layers.Dense(self.N_FEATURES, activation='sigmoid', name="ae_output")(x)
        self.autoencoder = keras.Model(inp, out, name="Autoencoder_MEC")
        self.autoencoder.compile(optimizer='adam', loss='mse')

    def _build_lstm(self):
        """
        Red LSTM que predice la salud futura del motor.
        Entrada: secuencia de SEQ_LEN pasos × N_FEATURES columnas
        Salida:  salud predicha en el siguiente tick (0-1)
        """
        inp = keras.Input(shape=(self.SEQ_LEN, self.N_FEATURES), name="lstm_input")
        x   = keras.layers.LSTM(32, return_sequences=True, name="lstm1")(inp)
        x   = keras.layers.Dropout(0.2)(x)
        x   = keras.layers.LSTM(16, name="lstm2")(x)
        out = keras.layers.Dense(1, activation='sigmoid', name="salud_pred")(x)
        self.lstm_model = keras.Model(inp, out, name="LSTM_Predictor_MEC")
        self.lstm_model.compile(optimizer='adam', loss='mse')

    def _features_to_array(self, features):
        return np.array(features, dtype=np.float32).reshape(1, -1)

    def update(self, features):
        """
        Llamado en cada tick lento. Actualiza buffers, entrena si hay suficientes
        datos, y devuelve (estado, mensaje, métricas_dict).
        """
        self.samples_seen += 1
        self.training_buf.append(features)

        # FASE 1: RECOLECCIÓN
        if len(self.training_buf) < self.MIN_SAMPLES:
            pct = (len(self.training_buf) / self.MIN_SAMPLES) * 100
            self.estado    = "APRENDIENDO"
            self.confianza = pct
            return self._make_result(
                f"⚙️ TF-IA: Recolectando datos base... {pct:.1f}% ({len(self.training_buf)}/{self.MIN_SAMPLES} muestras)"
            )

        # FASE 2: ENTRENAMIENTO INICIAL
        if not self.is_trained:
            self._train_initial()
            self.is_trained = True

        # FASE 3: INFERENCIA EN TIEMPO REAL
        return self._infer(features)

    def _train_initial(self):
        """Entrena Autoencoder, LSTM e IsolationForest con los datos recolectados."""
        data_raw = np.array(self.training_buf, dtype=np.float32)

        # Escalar datos 0-1
        data_scaled = self.scaler.fit_transform(data_raw)

        # Entrenar Autoencoder (sin salida en consola)
        self.autoencoder.fit(
            data_scaled, data_scaled,
            epochs=30, batch_size=16,
            verbose=0, shuffle=True
        )

        # Calibrar umbral con el error de reconstrucción del set de entrenamiento
        recon    = self.autoencoder.predict(data_scaled, verbose=0)
        losses   = np.mean(np.square(data_scaled - recon), axis=1)
        self.mean_loss    = float(np.mean(losses))
        self.std_loss     = float(np.std(losses)) + 1e-8
        self.threshold_ae = self.mean_loss + 3 * self.std_loss  # 3-sigma

        # Entrenar LSTM si hay suficientes secuencias
        if len(data_scaled) >= self.SEQ_LEN + 1:
            X_seq, y_seq = [], []
            for i in range(len(data_scaled) - self.SEQ_LEN):
                X_seq.append(data_scaled[i:i+self.SEQ_LEN])
                # Salud objetivo = feature 6 (índice del vib normalizado invertido)
                y_seq.append(1.0 - data_scaled[i+self.SEQ_LEN, 6])
            X_seq = np.array(X_seq); y_seq = np.array(y_seq)
            self.lstm_model.fit(X_seq, y_seq, epochs=20, batch_size=8, verbose=0)

        # Entrenar IsolationForest
        self.iso_forest.fit(data_scaled)

    def _infer(self, features):
        """Inferencia completa: AE + LSTM + IsolationForest."""
        raw = np.array(features, dtype=np.float32).reshape(1, -1)
        scaled = self.scaler.transform(raw)

        # Autoencoder: error de reconstrucción
        recon    = self.autoencoder.predict(scaled, verbose=0)
        ae_loss  = float(np.mean(np.square(scaled - recon)))
        self.ae_loss = ae_loss

        # Score normalizado: qué tan lejos está del umbral (0=normal, >1=anómalo)
        anomaly_score = (ae_loss - self.mean_loss) / self.std_loss
        self.anomaly_score = float(np.clip(anomaly_score, -3, 10))

        # LSTM: predicción de salud futura
        self.seq_buf.append(scaled[0])
        if len(self.seq_buf) > self.SEQ_LEN:
            self.seq_buf.pop(0)

        pred_salud = 100.0
        if len(self.seq_buf) == self.SEQ_LEN:
            seq_arr   = np.array(self.seq_buf, dtype=np.float32).reshape(1, self.SEQ_LEN, self.N_FEATURES)
            pred_raw  = float(self.lstm_model.predict(seq_arr, verbose=0)[0][0])
            pred_salud = np.clip(pred_raw * 100, 0, 100)
        self.pred_salud = pred_salud

        # IsolationForest: detección outlier
        iso_pred  = self.iso_forest.predict(scaled)[0]   # -1 = anomalía
        iso_score = float(self.iso_forest.decision_function(scaled)[0])

        # Decisión combinada
        ae_anomaly  = ae_loss > self.threshold_ae
        iso_anomaly = iso_pred == -1
        leve        = self.anomaly_score > 1.5

        # Historial
        self.hist_loss_ae.append(ae_loss)
        self.hist_pred_sal.append(pred_salud)
        self.hist_anomaly.append(float(np.clip(anomaly_score, 0, 5)))
        if len(self.hist_loss_ae) > 300:
            self.hist_loss_ae.pop(0)
            self.hist_pred_sal.pop(0)
            self.hist_anomaly.pop(0)

        self.confianza = 100.0

        if ae_anomaly and iso_anomaly:
            self.n_anomalias += 1
            self.estado = "ANOMALÍA CRÍTICA"
            msg = (f"🚨 TF-CRÍTICO: AE-Loss={ae_loss:.4f} (umbral={self.threshold_ae:.4f}) | "
                   f"IF=-1 | Salud predicha={pred_salud:.1f}% | Anomalías acum.={self.n_anomalias}")
        elif ae_anomaly:
            self.estado = "ANOMALÍA AE"
            msg = (f"⚠️ TF-AE: Error de reconstrucción elevado ({ae_loss:.4f}). "
                   f"La firma eléctrica se desvía del patrón aprendido.")
        elif iso_anomaly:
            self.estado = "ANOMALÍA IF"
            msg = (f"⚠️ TF-IF: IsolationForest detectó outlier (score={iso_score:.3f}). "
                   f"Salud predicha LSTM={pred_salud:.1f}%")
        elif leve:
            self.estado = "PREVENCIÓN"
            msg = (f"⚡ TF-AVISO: Desviación leve (σ={anomaly_score:.2f}). "
                   f"Salud predicha={pred_salud:.1f}%. Monitorear evolución.")
        else:
            self.estado = "NOMINAL"
            msg = (f"✅ TF-NOMINAL: AE-Loss={ae_loss:.5f} | σ={anomaly_score:.2f} | "
                   f"Salud LSTM predicha={pred_salud:.1f}%")

        return self._make_result(msg)

    def _make_result(self, msg):
        return {
            "estado":        self.estado,
            "mensaje":       msg,
            "confianza":     self.confianza,
            "pred_salud":    self.pred_salud,
            "anomaly_score": self.anomaly_score,
            "ae_loss":       self.ae_loss,
            "n_anomalias":   self.n_anomalias,
            "samples":       self.samples_seen,
            "hist_ae":       list(self.hist_loss_ae),
            "hist_sal":      list(self.hist_pred_sal),
            "hist_anom":     list(self.hist_anomaly),
        }

    def get_verificacion_matematica(self, v, i, p, q, s, pf, thd, vib, freq, temp, salud):
        """
        Verifica matemáticamente que todas las fórmulas eléctricas son coherentes.
        Devuelve lista de líneas para el log.
        """
        errores = []
        advertencias = []
        ok_list = []

        # 1. Verificar S = V * I
        s_calc = v * i
        err_s = abs(s_calc - s) / max(s, 1e-3) * 100
        if err_s < 2.0:
            ok_list.append(f"✅ S = V×I = {v:.2f}×{i:.2f} = {s_calc:.1f} VA (error={err_s:.2f}%)")
        else:
            advertencias.append(f"⚠️ S calculada={s_calc:.1f} VA ≠ S sistema={s:.1f} VA (error={err_s:.1f}%)")

        # 2. Verificar PF = P / S
        pf_calc = p / s if s > 0 else 0
        err_pf = abs(pf_calc - pf) * 100
        if err_pf < 2.0:
            ok_list.append(f"✅ PF = P/S = {p:.1f}/{s:.1f} = {pf_calc:.4f} (Δ={err_pf:.2f}%)")
        else:
            advertencias.append(f"⚠️ PF calculado={pf_calc:.4f} ≠ PF sistema={pf:.4f}")

        # 3. Verificar Q = sqrt(S²-P²)
        q_calc = np.sqrt(max(0, s**2 - p**2))
        err_q = abs(q_calc - q) / max(q, 1e-3) * 100
        if err_q < 2.0:
            ok_list.append(f"✅ Q = √(S²-P²) = {q_calc:.1f} VAr (error={err_q:.2f}%)")
        else:
            advertencias.append(f"⚠️ Q calculada={q_calc:.1f} VAr ≠ Q sistema={q:.1f} VAr")

        # 4. Triángulo de potencia: P²+Q²=S²
        lhs = p**2 + q**2
        rhs = s**2
        err_tri = abs(lhs - rhs) / max(rhs, 1e-3) * 100
        if err_tri < 2.0:
            ok_list.append(f"✅ Triángulo: P²+Q²≈S² (error={err_tri:.2f}%)")
        else:
            errores.append(f"❌ Triángulo INCOHERENTE: P²+Q²={lhs:.1f} ≠ S²={rhs:.1f}")

        # 5. Verificar Salud
        salud_calc = np.clip(1.0 - (vib/15.0)*0.7 - (thd/30.0)*0.3, 0, 1)
        err_sal = abs(salud_calc - salud) * 100
        if err_sal < 2.0:
            ok_list.append(f"✅ Salud = 1-(Vib/15)×0.7-(THD/30)×0.3 = {salud_calc*100:.1f}% (Δ={err_sal:.1f}%)")
        else:
            advertencias.append(f"⚠️ Salud calc={salud_calc*100:.1f}% ≠ Salud sistema={salud*100:.1f}%")

        # 6. Verificar rangos nominales
        if 110 <= v <= 140:
            ok_list.append(f"✅ Tensión {v:.1f}V en rango nominal [110-140V]")
        else:
            errores.append(f"❌ Tensión {v:.1f}V FUERA de rango nominal [110-140V]")

        if 55 <= freq <= 65 and freq > 0:
            ok_list.append(f"✅ Frecuencia {freq:.1f}Hz en rango nominal [55-65Hz]")
        elif freq > 0:
            errores.append(f"❌ Frecuencia {freq:.1f}Hz FUERA de rango [55-65Hz]")

        if thd < 5:
            ok_list.append(f"✅ THD={thd:.1f}% < 5% (IEEE 519 cumplido)")
        elif thd < 8:
            advertencias.append(f"⚠️ THD={thd:.1f}% entre 5-8% (límite aceptable)")
        else:
            errores.append(f"❌ THD={thd:.1f}% > 8% (VIOLA IEEE 519)")

        if pf >= 0.92:
            ok_list.append(f"✅ PF={pf:.3f} ≥ 0.92 (excelente)")
        elif pf >= 0.85:
            advertencias.append(f"⚠️ PF={pf:.3f} entre 0.85-0.92 (aceptable, mejorable)")
        else:
            errores.append(f"❌ PF={pf:.3f} < 0.85 (penalizable por CFE)")

        if vib < 2.8:
            ok_list.append(f"✅ Vibración {vib:.2f}mm/s < 2.8 (ISO 10816 Zona A)")
        elif vib < 7.1:
            advertencias.append(f"⚠️ Vibración {vib:.2f}mm/s en Zona B ISO 10816 (aceptable)")
        else:
            errores.append(f"❌ Vibración {vib:.2f}mm/s en Zona C/D ISO 10816 (PELIGROSO)")

        if temp < 50:
            ok_list.append(f"✅ Temperatura {temp:.1f}°C < 50°C (nominal)")
        elif temp < 65:
            advertencias.append(f"⚠️ Temperatura {temp:.1f}°C en zona de monitoreo [50-65°C]")
        else:
            errores.append(f"❌ Temperatura {temp:.1f}°C > 65°C (CRÍTICO)")

        resumen = "SIN ERRORES" if not errores else f"{len(errores)} ERROR(ES) DETECTADO(S)"
        return ok_list, advertencias, errores, resumen

# TELEMETRÍA DEL MOTOR
class MotorTelemetria:
    def __init__(self):
        self.sim_v    = 127.2
        self.sim_i    = 18.5
        self.sim_p    = 2213.0
        self.sim_vib  = 1.8
        self.sim_freq = 60.0
        self.sim_temp = 45.0
        self.sim_pf   = 0.88
        self.sim_thd  = 4.5
        self.falla    = False
        self.N        = 300

        self.hist_v   = np.full(self.N, self.sim_v,   dtype=float)
        self.hist_i   = np.full(self.N, self.sim_i,   dtype=float)
        self.hist_vib = np.full(self.N, self.sim_vib, dtype=float)
        self.fase     = 0.0
        self.t0       = time.time()

        v_init, i_init, p_init, q_init, s_init, pf_init, thd_init, vib_init, freq_init, temp_init, salud_init = self.get_metrics()
        self.current_v = v_init; self.current_i = i_init; self.current_p = p_init
        self.current_q = q_init; self.current_s = s_init; self.current_pf = pf_init
        self.current_thd = thd_init; self.current_vib = vib_init
        self.current_freq = freq_init; self.current_temp = temp_init
        self.current_salud = salud_init

        self.hist_rms   = np.zeros(self.N, dtype=float)
        self.hist_salud = np.full(500, salud_init, dtype=float)

        self.fast_N     = 3600
        self.fast_ts    = np.linspace(-60, 0, self.fast_N)
        self.fast_vib   = np.full(self.fast_N, vib_init, dtype=float)
        self.fast_thd   = np.full(self.fast_N, thd_init, dtype=float)
        self.fast_salud = np.full(self.fast_N, salud_init, dtype=float)
        self.ema_vib   = vib_init; self.ema_thd = thd_init; self.ema_salud = salud_init

        self.audit_v=[]; self.audit_i=[]; self.audit_p=[]; self.audit_vib=[]; self.audit_freq=[]
        self.audit_thd=[]; self.audit_temp=[]; self.audit_salud=[]; self.audit_ts=[]; self.audit_labels=[]
        self.audit_q=[]; self.audit_s=[]; self.audit_pf=[];  self.audit_espectro = []
        self.MAX_AUDIT = 120

        self.current_v_rms = 0.0
        self.current_i_rms = 0.0
        self.current_f_rms_global = 0.0

        self.N_samples = 16
        self.buffer_v = np.zeros(self.N_samples)
        self.fourier_mag, self.fourier_cos, self.fourier_sin = 0.0, 0.0, 0.0

    def aplicar_filtros_digitales(self, v_inst):
        self.buffer_v = np.roll(self.buffer_v, -1)
        self.buffer_v[-1] = v_inst
        k = np.arange(self.N_samples)
        theta = (2 * np.pi * k) / self.N_samples
        self.fourier_cos = (2.0 / self.N_samples) * np.sum(self.buffer_v * np.cos(theta))
        self.fourier_sin = (2.0 / self.N_samples) * np.sum(self.buffer_v * np.sin(theta))
        self.fourier_mag = np.sqrt(self.fourier_cos**2 + self.fourier_sin**2) / np.sqrt(2)

    def get_metrics(self):
        t = time.time()
        if self.falla:
            v    = np.random.uniform(40.0, 100.0)
            i    = np.random.uniform(50.0, 150.0)
            vib  = np.random.uniform(8.0, 25.0)
            temp = np.random.uniform(70.0, 115.0)
            freq = np.random.uniform(10.0, 180.0)
            thd  = np.random.uniform(15.0, 45.0)
        else:
            v    = self.sim_v    + np.sin(t*0.5)*1.2  + np.random.normal(0,0.08)
            i    = self.sim_i    - np.sin(t*0.5)*0.8  + np.random.normal(0,0.05)
            vib  = np.clip(self.sim_vib  + np.sin(t*0.4)*0.08 + np.random.normal(0,0.02), 0, 150)
            temp = np.clip(self.sim_temp + np.sin(t*0.3)*0.2, 0, 100)
            freq = max(0, self.sim_freq)
            thd  = np.clip(self.sim_thd + np.sin(t*0.6)*0.2 + np.random.normal(0,0.1), 0, 100)
        s = v * i
        p = s * np.clip(self.sim_pf + np.random.uniform(-0.005,0.005), 0, 1)
        pf = (p / s) if s > 0 else 0
        q = np.sqrt(max(0, s**2 - p**2))
        salud = np.clip(1.0 - (vib/15.0)*0.7 - (thd/30.0)*0.3, 0, 1)
        return v, i, p, q, s, pf, thd, vib, freq, temp, salud

    def tick(self, ext_data=None):
        if ext_data is not None:
            self.modo_externo = True
            v, i, vib, temp, freq, pf, thd, p, q, s, salud = ext_data
            salud = salud / 100.0
            self.ext_v=v; self.ext_i=i; self.ext_vib=vib; self.ext_temp=temp
            self.ext_freq=freq; self.ext_pf=pf; self.ext_thd=thd
            self.ext_p=p; self.ext_q=q; self.ext_s=s; self.ext_salud=salud

        if getattr(self, 'modo_externo', False):
            v,i,vib=self.ext_v,self.ext_i,self.ext_vib
            temp,freq,pf=self.ext_temp,self.ext_freq,self.ext_pf
            thd,p,q=self.ext_thd,self.ext_p,self.ext_q
            s,salud=self.ext_s,self.ext_salud
        else:
            v,i,p,q,s,pf,thd,vib,freq,temp,salud = self.get_metrics()

        self.current_v=v; self.current_i=i; self.current_p=p
        self.current_q=q; self.current_s=s; self.current_pf=pf
        self.current_thd=thd; self.current_vib=vib
        self.current_freq=freq; self.current_temp=temp; self.current_salud=salud

        self.hist_v   = np.append(self.hist_v[1:],   v)
        self.hist_i   = np.append(self.hist_i[1:],   i)
        self.hist_vib = np.append(self.hist_vib[1:], vib)
        self.hist_salud = np.append(self.hist_salud[1:], salud)
        self.fase -= 0.04

        alpha = 0.03
        self.ema_vib   = alpha*vib   + (1-alpha)*self.ema_vib
        self.ema_thd   = alpha*thd   + (1-alpha)*self.ema_thd
        self.ema_salud = alpha*salud + (1-alpha)*self.ema_salud
        ts = time.time() - self.t0
        self.fast_ts    = np.append(self.fast_ts[1:],    ts)
        self.fast_vib   = np.append(self.fast_vib[1:],   self.ema_vib)
        self.fast_thd   = np.append(self.fast_thd[1:],   self.ema_thd)
        self.fast_salud = np.append(self.fast_salud[1:], self.ema_salud)

        v_inst = (v*1.414)*np.sin(2*np.pi*60*time.time()) + np.random.normal(0, 2.0)
        self.aplicar_filtros_digitales(v_inst)
        return v,i,p,q,s,pf,thd,vib,freq,temp,salud

    def tick_audit(self, v,i,p,q,s,pf,vib,freq,thd,temp,salud):
        label = datetime.now().strftime("%I:%M:%S %p")
        espectro = [100.0, thd*1.2, thd*2.0, thd*0.8, thd*1.5, thd*0.5, thd*1.0, thd*0.4]
        self.audit_espectro.append(espectro)
        self.audit_labels.append(label); self.audit_v.append(v); self.audit_i.append(i)
        self.audit_p.append(p); self.audit_q.append(q); self.audit_s.append(s)
        self.audit_pf.append(pf); self.audit_vib.append(vib); self.audit_freq.append(freq)
        self.audit_thd.append(thd); self.audit_temp.append(temp); self.audit_salud.append(salud)
        for lst in [self.audit_labels,self.audit_v,self.audit_i,self.audit_p,self.audit_q,
                    self.audit_s,self.audit_pf,self.audit_vib,self.audit_freq,self.audit_thd,
                    self.audit_temp,self.audit_salud,self.audit_espectro]:
            if len(lst) > self.MAX_AUDIT: lst.pop(0)


# WIDGETS
class CardWidget(QWidget):
    def __init__(self, title, unit, color, bar_colors):
        super().__init__()
        self.color = color
        self.setMinimumWidth(140)
        self.setStyleSheet(f"background:{C_PANEL}; border:1px solid {C_MAIN}; border-radius:3px;")
        lay = QVBoxLayout(self); lay.setContentsMargins(12,8,12,8); lay.setSpacing(2)
        top = QHBoxLayout()
        lbl_t = QLabel(title.upper()); lbl_t.setStyleSheet(f"color:{C_TEXT_M}; font-size:9px; font-weight:bold; letter-spacing:0.5px; border:none;")
        lbl_u = QLabel(unit);  lbl_u.setStyleSheet(f"color:{C_NEUTRAL}; font-size:9px; border:none;")
        top.addWidget(lbl_t); top.addStretch(); top.addWidget(lbl_u)
        lay.addLayout(top)
        self.lbl_val = QLabel("--"); self.lbl_val.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.lbl_val.setStyleSheet(f"color:{C_TEXT_H}; font-family:'Consolas', monospace; font-size:24px; font-weight:normal; border:none;")
        lay.addWidget(self.lbl_val)
        self.bar_widget = BarWidget(bar_colors); self.bar_widget.setFixedHeight(4)
        lay.addWidget(self.bar_widget)

    def set_value(self, text, pct):
        self.lbl_val.setText(text); self.bar_widget.set_pct(pct)

class DualCardWidget(QWidget):
    def __init__(self, title, colors_pf, colors_thd):
        super().__init__()
        self.setMinimumWidth(180); self.setMinimumHeight(85)
        self.setStyleSheet(f"background:{C_PANEL}; border:1px solid {C_MAIN}; border-radius:3px;")
        lay = QVBoxLayout(self); lay.setContentsMargins(12,8,12,8); lay.setSpacing(2)
        lbl_title = QLabel(title.upper()); lbl_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_title.setStyleSheet(f"color:{C_TEXT_M}; font-size:9px; font-weight:bold; letter-spacing:0.5px; border:none;")
        lay.addWidget(lbl_title)
        val_lay = QHBoxLayout(); val_lay.setContentsMargins(0,0,0,0); val_lay.setSpacing(8)
        v_pf_lay = QVBoxLayout(); v_pf_lay.setSpacing(0)
        lbl_pf_title = QLabel("PF"); lbl_pf_title.setStyleSheet(f"color:{C_WHITE}; font-size:9px; border:none;"); v_pf_lay.addWidget(lbl_pf_title)
        self.lbl_pf_val = QLabel("0.00"); self.lbl_pf_val.setStyleSheet(f"color:{C_TEXT_H}; font-family:'Consolas', monospace; font-size:18px; border:none;")
        v_pf_lay.addWidget(self.lbl_pf_val)
        self.bar_pf = BarWidget(colors_pf); self.bar_pf.setFixedHeight(4); v_pf_lay.addWidget(self.bar_pf)
        v_thd_lay = QVBoxLayout(); v_thd_lay.setSpacing(0)
        lbl_thd_title = QLabel("% THD"); lbl_thd_title.setStyleSheet(f"color:{C_THD}; font-size:9px; border:none;"); lbl_thd_title.setAlignment(Qt.AlignmentFlag.AlignRight); v_thd_lay.addWidget(lbl_thd_title)
        self.lbl_thd_val = QLabel("0.0"); self.lbl_thd_val.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.lbl_thd_val.setStyleSheet(f"color:{C_TEXT_H}; font-family:'Consolas', monospace; font-size:18px; border:none;")
        v_thd_lay.addWidget(self.lbl_thd_val)
        self.bar_thd = BarWidget(colors_thd); self.bar_thd.setFixedHeight(4); v_thd_lay.addWidget(self.bar_thd)
        val_lay.addLayout(v_pf_lay); val_lay.addLayout(v_thd_lay)
        lay.addLayout(val_lay)

    def set_values(self, pf_text, pf_pct, thd_text, thd_pct):
        self.lbl_pf_val.setText(pf_text); self.bar_pf.set_pct(pf_pct)
        self.lbl_thd_val.setText(thd_text); self.bar_thd.set_pct(thd_pct)

class BarWidget(QWidget):
    def __init__(self, colors):
        super().__init__()
        self.colors = colors; self.pct = 0.5
        self.threshold_1 = 0.6; self.threshold_2 = 0.8
    def set_pct(self, p):
        self.pct = np.clip(p, 0, 1); self.update()
    def paintEvent(self, e):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C_BORDER))
        n = 30; seg_w = w / n; fill_limit_segment = int(self.pct * n)
        for i in range(n):
            seg_pct_pos = (i + 0.5) / n
            if seg_pct_pos <= self.threshold_1: color = QColor(self.colors[0])
            elif seg_pct_pos <= self.threshold_2: color = QColor(self.colors[1])
            else: color = QColor(self.colors[2])
            if i < fill_limit_segment:
                p.fillRect(int(i*seg_w), 0, int(max(1, seg_w-1)), h, color)


# VENTANA PRINCIPAL
class ConsolaCFE(QMainWindow):
    def __init__(self):
        super().__init__()
        print("¡El código está iniciando con TensorFlow integrado!")
        self.setWindowTitle("MEC Industrial Analytics Pro — Núcleo MEC")
        self.resize(1200, 750)
        self.setMinimumSize(1100, 700)
        self._center_window()

        self.motor     = MotorTelemetria()
        self.tf_brain  = TensorFlowBrain()
        self._tf_entrenando = False
        self.MAX_THD   = 30.0
        self.activo_nombre = "TIEMPO REAL"   # el mapa 3D lo cambia al activo seleccionado

        # Último resultado TF (para el log)
        self.tf_result = self.tf_brain._make_result("> Inicializando TensorFlow AI Engine...")

        self.serial_port  = None
        self.buffer_serial = ""
        self._conn_state  = 'connecting'

        try:
            self.serial_port = serial.Serial('COM11', 115200, timeout=0)
            print("🔗 Escuchando telemetría en COM11...")
        except Exception as e:
            print(f"⚠️ Enlace COM11 no disponible ({e}). Usando fuente de datos en vivo interna.")

        self._build_ui()

        if self.serial_port and self.serial_port.is_open:
            self._set_conn_style('connecting')
        else:
            self._set_conn_style('connected')

        self.timer = QTimer(); self.timer.timeout.connect(self._loop_fast); self.timer.start(16)
        self.timer_slow = QTimer(); self.timer_slow.timeout.connect(self._loop_slow); self.timer_slow.start(1000)

        self.watcher = QFileSystemWatcher([os.path.abspath(__file__)])
        self.watcher.fileChanged.connect(self._restart_app)

    def _center_window(self):
        qr = self.frameGeometry()
        cp = self.screen().availableGeometry().center()
        qr.moveCenter(cp); self.move(qr.topLeft())

    def _restart_app(self):
        print("🚀 Cambio detectado! Reiniciando...")
        script_path = os.path.abspath(sys.argv[0])
        subprocess.Popen(f'"{sys.executable}" "{script_path}"', shell=True)
        QApplication.quit(); sys.exit(0)

    def _init_smart_zoom(self, plt, name):
        setattr(self, f"zoomed_{name}", False)
        setattr(self, f"auto_{name}", False)
        timer = QTimer(); timer.setSingleShot(True)
        def reset_zoom(): setattr(self, f"zoomed_{name}", False)
        timer.timeout.connect(reset_zoom)
        setattr(self, f"timer_{name}", timer)
        def on_range_changed(*args):
            if getattr(self, f"auto_{name}", False): return
            setattr(self, f"zoomed_{name}", True)
            getattr(self, f"timer_{name}").start(5000)
        plt.getViewBox().sigRangeChanged.connect(on_range_changed)

    # UI BUILDER
    def _build_ui(self):
        root = QWidget(); self.setCentralWidget(root)
        root.setStyleSheet(STYLE_MAIN)
        main_lay = QVBoxLayout(root); main_lay.setContentsMargins(0,0,0,0); main_lay.setSpacing(0)
        main_lay.addWidget(self._build_topbar())
        body = QHBoxLayout(); body.setContentsMargins(0,0,0,0); body.setSpacing(0)
        body.addWidget(self._build_sidebar())
        body.addWidget(self._build_center(), 1)
        wbody = QWidget(); wbody.setLayout(body); main_lay.addWidget(wbody, 1)

    def _build_topbar(self):
        bar = QWidget(); bar.setFixedHeight(44)
        bar.setStyleSheet(
            f"background:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #0a1a28, stop:1 {C_PANEL});"
            f"border-bottom:1px solid {C_HUD_DK};"
        )
        lay = QHBoxLayout(bar); lay.setContentsMargins(16,0,16,0)
        lbl_logo  = QLabel("CFE"); lbl_logo.setStyleSheet(f"color:{C_GOOD}; font-size:16px; font-weight:900; font-style:italic; letter-spacing:1px;")
        lbl_mec = QLabel("◤ M · E · C")
        lbl_mec.setStyleSheet(f"color:{C_GLOW}; font-family:'Consolas',monospace; font-size:16px; font-weight:900; letter-spacing:3px;")
        lbl_title = QLabel("INDUSTRIAL ANALYTICS // NÚCLEO MEC")
        lbl_title.setStyleSheet(f"color:{C_HUD}; font-size:11px; font-weight:600; letter-spacing:2px;")
        self.lbl_tf_estado = QLabel("◈ MEC: INICIALIZANDO")
        self.lbl_tf_estado.setStyleSheet(f"color:{C_ACCENT}; font-size:10px; font-weight:bold; letter-spacing:1px;")
        dot_g = QLabel("● EN LÍNEA"); dot_g.setStyleSheet(f"color:{C_GOOD}; font-size:10px; font-weight:bold; letter-spacing:1px;")
        lay.addWidget(lbl_logo); lay.addSpacing(10); lay.addWidget(lbl_mec); lay.addSpacing(10)
        lay.addWidget(lbl_title); lay.addStretch()
        lay.addWidget(self.lbl_tf_estado); lay.addSpacing(20); lay.addWidget(dot_g)
        return bar

    def _build_sidebar(self):
        side = QWidget(); side.setFixedWidth(213)
        side.setStyleSheet(f"background:#0a121c; border-right:1px solid {C_HUD_DK};")
        lay = QVBoxLayout(side); lay.setContentsMargins(8,10,8,10); lay.setSpacing(6)

        self.lbl_conn_estado = QLabel("● ENLAZANDO…")
        self.lbl_conn_estado.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_conn_estado.setWordWrap(True); lay.addWidget(self.lbl_conn_estado)

        self.lbl_conn_detalle = QLabel("")
        self.lbl_conn_detalle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_conn_detalle.setWordWrap(True)
        self.lbl_conn_detalle.setStyleSheet("color:#7c8da0; font-size:10px; border:none;")
        lay.addWidget(self.lbl_conn_detalle)

        sep1 = QFrame(); sep1.setFrameShape(QFrame.Shape.HLine)
        sep1.setStyleSheet(f"color:{C_BORDER};"); lay.addWidget(sep1)

        # Núcleo MEC (KPIs del motor de inferencia)
        self.lbl_tf_info = QLabel("◤ MEC\nInicializando...")
        self.lbl_tf_info.setAlignment(Qt.AlignmentFlag.AlignLeft)
        self.lbl_tf_info.setWordWrap(True)
        self.lbl_tf_info.setStyleSheet(f"""
            QLabel {{
                color: {C_GLOW}; font-family: 'Consolas', monospace; font-size: 10px;
                border: 1px solid {C_HUD_DK}; border-radius: 4px; padding: 8px;
                background: {C_HUD_BG}; line-height: 1.5;
            }}
        """)
        lay.addWidget(self.lbl_tf_info)

        # LOG DEL SISTEMA (lo escribe MEC) — reubicado aquí desde la esquina inferior derecha
        lay.addWidget(self._lbl("MEC // Log del Sistema", C_GLOW))
        log_style = (f"background:{C_HUD_BG}; color:#E5F6FF; border:none; border-left:2px solid {{color}};"
                     f"font-family:'Consolas','Courier New',monospace; font-size:10px; padding:6px;")

        self.txt_sec_estado = QTextEdit(); self.txt_sec_estado.setReadOnly(True)
        self.txt_sec_estado.setStyleSheet(log_style.format(color=C_HUD))
        self.txt_sec_estado.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.txt_sec_estado.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        lay.addWidget(self.txt_sec_estado, 3)

        self.txt_sec_analisis = QTextEdit(); self.txt_sec_analisis.setReadOnly(True)
        self.txt_sec_analisis.setStyleSheet(log_style.format(color=C_GOOD))
        self.txt_sec_analisis.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.txt_sec_analisis.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        lay.addWidget(self.txt_sec_analisis, 4)

        # Alias para compatibilidad con el resto del código
        self.txt_reporte = self.txt_sec_analisis
        self.txt_sec_metricas = self.txt_sec_estado

        btn_excel = QPushButton("📊  EXPORTAR A EXCEL")
        btn_excel.setStyleSheet(f"""
            QPushButton {{ background:#003820; border:1px solid {C_GOOD}; color:{C_GOOD}; font-size:10px; font-weight:bold; padding:7px 4px; border-radius:4px; }}
            QPushButton:hover {{ background:{C_GOOD}; color:#000; font-weight:bold; }}
        """)
        btn_excel.clicked.connect(self._exportar_excel); lay.addWidget(btn_excel)
        self.lbl_excel_status = QLabel(""); self.lbl_excel_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_excel_status.setWordWrap(True)
        self.lbl_excel_status.setStyleSheet("color:#7c8da0; font-size:9px; border:none;")
        lay.addWidget(self.lbl_excel_status)
        return side

    def _set_conn_style(self, state):
        self._conn_state = state
        if state == 'connecting':
            color, texto, bg, borde = C_HUD, "● ENLAZANDO…", "#06283d", C_HUD_DK
            detalle = "Puerto COM11 abierto\nEsperando telemetría en vivo"
        elif state == 'waiting':
            color, texto, bg, borde = C_ACCENT, "● ESPERANDO DATOS", "#3a2e00", "#7A4A00"
            detalle = "Reintentando enlace de datos…"
        else:  # connected
            activo = getattr(self, 'activo_nombre', 'TIEMPO REAL')
            color, texto, bg, borde = C_GOOD, f"● CONECTADO · {activo.upper()}", "#00331A", "#006633"
            detalle = "Telemetría en vivo del activo"
        if hasattr(self, 'lbl_conn_estado'):
            self.lbl_conn_estado.setText(texto)
            self.lbl_conn_estado.setStyleSheet(
                f"color:{color}; background:{bg}; border:1px solid {borde}; border-radius:3px; "
                f"font-size:10px; font-weight:bold; padding:6px; letter-spacing:1px;"
            )
        if hasattr(self, 'lbl_conn_detalle'):
            self.lbl_conn_detalle.setText(detalle)

    def set_activo(self, nombre: str):
        """El mapa 3D llama a esto al seleccionar una estructura de CFE
        (subestación, oficina, generador, torre…). Cambia la etiqueta de conexión."""
        self.activo_nombre = nombre or "TIEMPO REAL"
        self._set_conn_style('connected')

    def _build_center(self):
        center = QWidget(); lay = QVBoxLayout(center); lay.setContentsMargins(4,4,4,4); lay.setSpacing(6)
        lay.addWidget(self._build_cards())

        # FILA 1: Triángulo, Oscilograma, Tendencias (3 paneles)
        row1 = QWidget(); r1 = QHBoxLayout(row1); r1.setContentsMargins(0,0,0,0); r1.setSpacing(6)
        r1.addWidget(self._build_triangle_panel(), 1)
        r1.addWidget(self._build_oscilo_panel(), 2)
        r1.addWidget(self._build_stacked_panel(), 1)
        lay.addWidget(row1, 1)

        # FILA 2: f_RMS Global, Monitor de Condición, Espectro Armónico (3 paneles)
        row2 = QWidget(); r2 = QHBoxLayout(row2); r2.setContentsMargins(0,0,0,0); r2.setSpacing(6)
        r2.addWidget(self._build_frms_panel(), 2)
        r2.addWidget(self._build_auditoria_panel(), 2)
        r2.addWidget(self._build_espectro_panel(), 1)
        lay.addWidget(row2, 1)
        return center

    def _build_cards(self):
        w = QWidget(); w.setFixedHeight(85); lay = QHBoxLayout(w); lay.setContentsMargins(0,0,0,0); lay.setSpacing(6)
        standard_colors = [C_GOOD, C_ACCENT, C_CRIT]
        pf_colors = [C_CRIT, C_ACCENT, C_GOOD]
        thd_colors = [C_GOOD, C_ACCENT, C_CRIT]
        self.card_v   = CardWidget("Tensión RMS",    "V",     C_VOLT, standard_colors)
        self.card_i   = CardWidget("Corriente RMS",  "A",     C_CURR, standard_colors)
        self.card_p   = CardWidget("Potencia Activa","W",     C_P_ACT, standard_colors)
        self.card_vib = CardWidget("Vibración Global","mm/s", C_VIB,  standard_colors)
        self.card_t   = CardWidget("Temp Estator",   "°C",   C_TEMP, standard_colors)
        self.card_pf  = DualCardWidget("Calidad de Red", pf_colors, thd_colors)
        for c in [self.card_v, self.card_i, self.card_p, self.card_vib, self.card_t, self.card_pf]: lay.addWidget(c)
        return w

    def _build_frms_panel(self):
        w = QWidget(); w.setStyleSheet(f"background:{C_PANEL}; border:1px solid {C_MAIN}; border-radius:4px;")
        lay = QVBoxLayout(w); lay.setContentsMargins(10,10,10,10); lay.setSpacing(6); lay.addWidget(self._lbl("Análisis f_RMS Global", C_TEXT_H))
        self.plt_rms = pg.PlotWidget(); self.plt_rms.setBackground(C_PANEL); self.plt_rms.setMinimumHeight(90)
        self.plt_rms.showGrid(x=True, y=True, alpha=0.15)
        self.plt_rms.getAxis('left').setTextPen(C_NEUTRAL); self.plt_rms.getAxis('bottom').setTextPen(C_NEUTRAL)
        self.plt_rms.getAxis('left').setPen(C_BORDER); self.plt_rms.getAxis('bottom').setPen(C_BORDER)
        self.line_rms = self.plt_rms.plot(pen=_holo_pen(C_RMS, 2), shadowPen=_glow_pen(C_RMS, 8),
                                          fillLevel=0, brush=_area_brush(C_RMS, 45)); self._setup_time_axis(self.plt_rms)
        self._init_smart_zoom(self.plt_rms, 'rms'); lay.addWidget(self.plt_rms)
        self.tbl_rms = self._make_table(["Hora","F_RMS","V","I","P","Q","S","PF","THD","Vib","Freq","Temp"], 12); lay.addWidget(self.tbl_rms); return w

    def _build_oscilo_panel(self):
        w = QWidget(); w.setStyleSheet(f"background:{C_PANEL}; border:1px solid {C_MAIN}; border-radius:4px;")
        lay = QVBoxLayout(w); lay.setContentsMargins(10,10,10,10); lay.setSpacing(6); lay.addWidget(self._lbl("Oscilograma", C_TEXT_H))
        self.plt_osc = pg.PlotWidget(); self.plt_osc.setBackground(C_PANEL)
        self.plt_osc.showGrid(x=True, y=True, alpha=0.15)
        self.plt_osc.getAxis('left').setTextPen(C_NEUTRAL); self.plt_osc.getAxis('bottom').setTextPen(C_NEUTRAL)
        self.plt_osc.getAxis('left').setPen(C_BORDER); self.plt_osc.getAxis('bottom').setPen(C_BORDER)
        leyenda_osc = self.plt_osc.addLegend(offset=(5,-5))
        leyenda_osc.setBrush(pg.mkBrush('#00000000')); leyenda_osc.setPen(pg.mkPen('#00000000'))
        self.line_ov = self.plt_osc.plot(pen=_holo_pen(C_VOLT, 2.2), shadowPen=_glow_pen(C_VOLT, 7), name="Tensión (V)")
        self.line_oi = self.plt_osc.plot(pen=_holo_pen(C_CURR, 2.2), shadowPen=_glow_pen(C_CURR, 7), name="Corriente (A)")
        self._init_smart_zoom(self.plt_osc, 'osc'); lay.addWidget(self.plt_osc)
        self.tbl_osc = self._make_table(["Hora","Tensión","Corriente","Potencia","Frecuencia"], 5); lay.addWidget(self.tbl_osc); return w

    def _build_espectro_panel(self):
        w = QWidget(); w.setStyleSheet(f"background:{C_PANEL}; border:1px solid {C_MAIN}; border-radius:4px;")
        lay = QVBoxLayout(w); lay.setContentsMargins(10,10,10,10); lay.setSpacing(6); lay.addWidget(self._lbl("Espectro Armónico (THD)", C_TEXT_H))
        self.plt_espectro = pg.PlotWidget(); self.plt_espectro.setBackground(C_PANEL)
        self.plt_espectro.showGrid(x=False, y=True, alpha=0.15)
        self.plt_espectro.getAxis('left').setTextPen(C_NEUTRAL); self.plt_espectro.getAxis('bottom').setTextPen(C_NEUTRAL)
        self.plt_espectro.getAxis('left').setPen(C_BORDER); self.plt_espectro.getAxis('bottom').setPen(C_BORDER)
        self.x_harms = [1,2,3,4,5,6,7,8]
        self.bar_espectro = pg.BarGraphItem(x=self.x_harms, height=[0]*8, width=0.55, brush=pg.mkBrush(C_THD), pen=pg.mkPen(C_GLOW, width=1.2))
        self.plt_espectro.addItem(self.bar_espectro); self.plt_espectro.setYRange(0, 110)
        self.plt_espectro.getAxis('bottom').setTicks([[(h, str(h)) for h in self.x_harms]])
        self._init_smart_zoom(self.plt_espectro, 'esp'); lay.addWidget(self.plt_espectro)
        self.tbl_espectro = self._make_table(["Armónico","Mag (%)","THD Cont."], 3); lay.addWidget(self.tbl_espectro); return w

    def _build_stacked_panel(self):
        w = QWidget(); w.setStyleSheet(f"background:{C_PANEL}; border:1px solid {C_MAIN}; border-radius:4px;")
        lay = QVBoxLayout(w); lay.setContentsMargins(10,10,10,10); lay.setSpacing(6); lay.addWidget(self._lbl("Tendencias Históricas", C_TEXT_H))
        for title, col, attr, z_name in [("Tensión (V)", C_VOLT, "plt_st_v", "st_v"),
                                          ("Corriente (A)", C_CURR, "plt_st_i", "st_i"),
                                          ("Vibración Mecánica", C_VIB, "plt_st_vib", "st_vib")]:
            lbl = QLabel(title); lbl.setStyleSheet(f"color:{C_TEXT_M}; font-size:9px; text-transform:uppercase;")
            plt = pg.PlotWidget(); plt.setBackground(C_PANEL); plt.setFixedHeight(60)
            plt.showGrid(x=False, y=True, alpha=0.1)
            plt.getAxis('left').setTextPen(C_NEUTRAL); plt.getAxis('left').setPen(C_BORDER); plt.getAxis('left').setWidth(40)
            plt.getAxis('bottom').setTextPen(C_NEUTRAL); plt.getAxis('bottom').setPen(C_BORDER)
            line = plt.plot(pen=_holo_pen(col, 2), shadowPen=_glow_pen(col, 6), fillLevel=0, brush=_area_brush(col, 26))
            self._init_smart_zoom(plt, z_name); lay.addWidget(lbl); lay.addWidget(plt)
            setattr(self, attr, (plt, line))
        return w

    def _build_triangle_panel(self):
        w = QWidget(); w.setStyleSheet(f"background:{C_PANEL}; border:1px solid {C_MAIN}; border-radius:4px;"); w.setMinimumWidth(220)
        lay = QVBoxLayout(w); lay.setContentsMargins(10,10,10,10); lay.setSpacing(6); lay.addWidget(self._lbl("Triángulo de Potencia", C_TEXT_H))
        self.plt_tri = pg.PlotWidget(); self.plt_tri.setBackground(C_PANEL)
        self.plt_tri.showGrid(x=True, y=True, alpha=0.1)
        self.plt_tri.getAxis('left').setTextPen(C_NEUTRAL); self.plt_tri.getAxis('bottom').setTextPen(C_NEUTRAL)
        self.plt_tri.getAxis('left').setPen(C_BORDER); self.plt_tri.getAxis('bottom').setPen(C_BORDER)
        self.ln_p = self.plt_tri.plot(pen=_holo_pen(C_P_ACT, 2.5), shadowPen=_glow_pen(C_P_ACT, 7))
        self.ln_q = self.plt_tri.plot(pen=_holo_pen(C_P_REA, 2.5), shadowPen=_glow_pen(C_P_REA, 7))
        self.ln_s = self.plt_tri.plot(pen=pg.mkPen(C_P_APP, width=1.4, style=Qt.PenStyle.DashLine))
        self.txt_P = pg.TextItem("P", color=C_P_ACT, anchor=(0.5,0.5)); self.plt_tri.addItem(self.txt_P)
        self.txt_Q = pg.TextItem("Q", color=C_P_REA, anchor=(0.5,0.5)); self.plt_tri.addItem(self.txt_Q)
        self.txt_S = pg.TextItem("S", color=C_P_APP, anchor=(0.5,0.5)); self.plt_tri.addItem(self.txt_S)
        self._init_smart_zoom(self.plt_tri, 'tri'); lay.addWidget(self.plt_tri)
        self.tbl_tri = self._make_table(["Hora","V","Salud","THD"], 4); lay.addWidget(self.tbl_tri); return w

    def _build_auditoria_panel(self):
        w = QWidget(); w.setStyleSheet(f"background:{C_PANEL}; border:1px solid {C_MAIN}; border-radius:4px;")
        lay = QVBoxLayout(w); lay.setContentsMargins(10,10,10,10); lay.setSpacing(6)
        lay.addWidget(self._lbl("Monitor de Condición", C_TEXT_H))
        self.plt_audit = pg.PlotWidget(); self.plt_audit.setBackground(C_PANEL); self.plt_audit.setMinimumHeight(90)
        self.plt_audit.showGrid(x=True, y=True, alpha=0.15)
        self.plt_audit.getAxis('left').setTextPen(C_NEUTRAL); self.plt_audit.getAxis('bottom').setTextPen(C_NEUTRAL)
        self.plt_audit.getAxis('left').setPen(C_BORDER); self.plt_audit.getAxis('bottom').setPen(C_BORDER)
        self.ln_aud_salud = pg.PlotCurveItem(pen=_holo_pen(C_SALUD, 2)); self.ln_aud_salud.setShadowPen(_glow_pen(C_SALUD, 6)); self.plt_audit.addItem(self.ln_aud_salud)
        self.ln_aud_thd   = pg.PlotCurveItem(pen=_holo_pen(C_THD, 2));   self.ln_aud_thd.setShadowPen(_glow_pen(C_THD, 6));   self.plt_audit.addItem(self.ln_aud_thd)
        self.ln_aud_vib   = pg.PlotCurveItem(pen=_holo_pen(C_VIB, 2));   self.ln_aud_vib.setShadowPen(_glow_pen(C_VIB, 6));   self.plt_audit.addItem(self.ln_aud_vib)
        leyenda_aud = self.plt_audit.addLegend(offset=(5,-5))
        leyenda_aud.setBrush(pg.mkBrush('#00000000')); leyenda_aud.setPen(pg.mkPen('#00000000'))
        leyenda_aud.addItem(self.ln_aud_salud, "Salud (%)"); leyenda_aud.addItem(self.ln_aud_thd, "THD (%)"); leyenda_aud.addItem(self.ln_aud_vib, "Vibración (mm/s)")
        self._init_smart_zoom(self.plt_audit, 'audit'); lay.addWidget(self.plt_audit)
        self.tbl_audit = self._make_table(["Hora","V RMS","I RMS","P","Vib","Freq","THD","Temp","Salud"], 9)
        lay.addWidget(self.tbl_audit); return w
    
    def _build_reporte_ia(self):
        w = QWidget(); w.setStyleSheet(f"background:{C_PANEL}; border:1px solid {C_MAIN}; border-radius:4px;")
        lay = QVBoxLayout(w); lay.setContentsMargins(10,10,10,10); lay.setSpacing(4)
        lay.addWidget(self._lbl("MEC // Log del Sistema", C_GLOW))

        # Estilo base
        txt_style = (f"background:#0B0F19; color:{C_WHITE}; border:none; border-left:2px solid {{color}};"
                     f"border-radius:0px; font-family:'Consolas','Courier New',monospace; font-size:10px; padding:6px;")

        # Sección 1: Estado + Métricas
        self.txt_sec_estado = QTextEdit(); self.txt_sec_estado.setReadOnly(True)
        self.txt_sec_estado.setFixedHeight(155) 
        self.txt_sec_estado.setStyleSheet(txt_style.format(color=C_HUD))
        
        # ELIMINAMOS SCROLLS AQUÍ
        self.txt_sec_estado.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.txt_sec_estado.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        lay.addWidget(self.txt_sec_estado)

        # Sección 2: Análisis IA + Recomendaciones Fusionadas
        self.txt_sec_analisis = QTextEdit(); self.txt_sec_analisis.setReadOnly(True)
        self.txt_sec_analisis.setFixedHeight(160) 
        self.txt_sec_analisis.setStyleSheet(txt_style.format(color=C_GOOD))
        
        # ELIMINAMOS SCROLLS AQUÍ
        self.txt_sec_analisis.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.txt_sec_analisis.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        lay.addWidget(self.txt_sec_analisis)

        # Alias para compatibilidad
        self.txt_reporte = self.txt_sec_analisis
        self.txt_sec_metricas = self.txt_sec_estado
        return w

    def _lbl(self, text, color=C_TEXT_M, size=11):
        l = QLabel("▌ " + text.upper())
        l.setStyleSheet(f"color:{color}; font-size:{size}px; font-weight:bold; letter-spacing:1px; border:none;")
        return l

    def _mec_html(self, titulo, cuerpo, accent):
        """Renderiza un bloque del log con estética HUD futurista, firmado por MEC."""
        import html as _html
        cuerpo_esc = _html.escape(cuerpo)
        return (
            "<div style=\"font-family:'Consolas',monospace;\">"
            f"<div style='color:{accent}; font-weight:bold; letter-spacing:2px; "
            f"border-bottom:1px solid {accent}; padding-bottom:3px; margin-bottom:5px;'>"
            f"&#9670; {titulo} <span style='color:{C_HUD};'>&#9646;</span></div>"
            f"<pre style='color:#E5F6FF; font-family:Consolas,monospace; font-size:10px; "
            f"white-space:pre-wrap; margin:0; line-height:1.45;'>{cuerpo_esc}</pre>"
            "</div>"
        )

    def _make_table(self, headers, cols):
        num_filas = 6
        t = QTableWidget(num_filas, cols)
        t.setHorizontalHeaderLabels(headers)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        t.verticalHeader().setVisible(False)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        t.setAlternatingRowColors(True)
        t.setFixedHeight(num_filas * 18 + 26)
        for r in range(num_filas):
            for c in range(cols):
                item = QTableWidgetItem("--"); item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                t.setItem(r, c, item); t.setRowHeight(r, 18)
        return t

    def _fill_table(self, tbl, data):
        for r, row in enumerate(data[:tbl.rowCount()]):
            for c, val in enumerate(row[:tbl.columnCount()]):
                item = tbl.item(r, c)
                if item: item.setText(str(val)); item.setForeground(QColor(C_WHITE))

    def _setup_time_axis(self, plt):
        plt.getAxis('bottom').setLabel('', color=C_NEUTRAL)

    def _smooth(self, data, alpha=0.15):
        if len(data) == 0:
            return data
        result = np.array(data, dtype=float)
        for i in range(1, len(result)):
            result[i] = alpha * result[i] + (1 - alpha) * result[i - 1]
        return result

    # BUCLE RÁPIDO (16ms)
    def _loop_fast(self):
        ext_data = None
        if self.serial_port and self.serial_port.is_open:
            try:
                if self.serial_port.in_waiting > 0:
                    raw_data = self.serial_port.read(self.serial_port.in_waiting).decode('utf-8', errors='ignore')
                    self.buffer_serial += raw_data
                    if '\n' in self.buffer_serial:
                        lineas = self.buffer_serial.split('\n'); self.buffer_serial = lineas[-1]
                        for linea in reversed(lineas[:-1]):
                            linea = linea.strip()
                            if linea:
                                valores = linea.split(',')
                                if len(valores) == 11:
                                    ext_data = [float(x) for x in valores]
                                    if self._conn_state != 'connected': self._set_conn_style('connected')
                                    break
            except Exception: pass

        v_base,i_base,p,q,s,pf,thd,vib,freq,temp,salud = self.motor.tick(ext_data=ext_data)

        samples = 200
        if freq > 0:
            t_w = np.linspace(0, 1/freq, samples)
            v_t = (v_base*1.414)*np.sin(2*np.pi*freq*t_w)
            i_t = (i_base*1.414)*np.sin(2*np.pi*freq*t_w - np.arccos(min(max(pf,-1),1)))
        else:
            t_w = np.linspace(0, 0.1, samples)
            v_t = np.full(samples, v_base) + np.random.normal(0, v_base*0.002, samples)
            i_t = np.full(samples, i_base) + np.random.normal(0, i_base*0.002, samples)

        if freq > 0:
            v_rms_calc = np.sqrt(np.mean(np.square(v_t))); i_rms_calc = np.sqrt(np.mean(np.square(i_t)))
        else:
            v_rms_calc = np.mean(np.abs(v_t)); i_rms_calc = np.mean(np.abs(i_t))
        f_rms_global = np.sqrt(v_rms_calc * i_rms_calc)

        self.motor.hist_rms = np.append(self.motor.hist_rms[1:], f_rms_global)
        x_rms = np.arange(len(self.motor.hist_rms)); self.line_rms.setData(x_rms, self.motor.hist_rms)
        self.motor.current_v_rms = v_rms_calc; self.motor.current_i_rms = i_rms_calc; self.motor.current_f_rms_global = f_rms_global

        self.card_v.set_value(f"{v_rms_calc:.1f}", v_rms_calc/160)
        self.card_i.set_value(f"{i_rms_calc:.2f}", i_rms_calc/100)
        self.card_p.set_value(f"{p:.0f}", p/15000)
        self.card_vib.set_value(f"{vib:.2f}", vib/15)
        self.card_t.set_value(f"{temp:.1f}", temp/100)
        scaled_thd = (thd/self.MAX_THD) if thd > 0 else 0
        self.card_pf.set_values(f"{pf:.2f}", pf, f"{thd:.1f}%", scaled_thd)

        datos_validos = self.motor.hist_rms[self.motor.hist_rms > 0]
        if len(datos_validos) > 1:
            y_centro = np.mean(datos_validos); y_rango = max(y_centro*0.8, 100)
            y_min = max(0, y_centro-y_rango); y_max = y_centro+y_rango
        else:
            y_min=0; y_max=200
        self.auto_rms = True
        self.plt_rms.getViewBox().setLimits(xMin=0, xMax=len(x_rms), yMin=y_min, yMax=y_max)
        if not self.zoomed_rms:
            self.plt_rms.setXRange(0, len(x_rms), padding=0); self.plt_rms.setYRange(y_min, y_max, padding=0)
        self.auto_rms = False

        if freq > 0:
            t_w2 = np.linspace(0, (1/freq)*2.2, 200)
            vw = (v_base*1.414)*np.sin(2*np.pi*freq*t_w2 - self.motor.fase)
            iw = (i_base*1.414)*np.sin(2*np.pi*freq*t_w2 - np.arccos(min(max(pf,-1),1)) - self.motor.fase)
            x_max_osc = (1/freq)*2.2*1000
        else:
            t_w2 = np.linspace(0, 0.1, 200); vw = np.full(200, v_base)+np.random.normal(0, v_base*0.002, 200)
            iw = np.full(200, i_base)+np.random.normal(0, i_base*0.002, 200); x_max_osc = 100.0
        self.line_ov.setData(t_w2*1000, vw); self.line_oi.setData(t_w2*1000, iw); self.auto_osc = True
        y_max_osc = max(np.max(np.abs(vw)), np.max(np.abs(iw)), 220)
        self.plt_osc.getViewBox().setLimits(xMin=0, xMax=x_max_osc, yMin=-y_max_osc, yMax=y_max_osc)
        if not self.zoomed_osc:
            self.plt_osc.setXRange(0, x_max_osc, padding=0); self.plt_osc.setYRange(-y_max_osc, y_max_osc, padding=0)
        self.auto_osc = False

        h_vals = [100.0, thd*1.2, thd*2.0, thd*0.8, thd*1.5, thd*0.5, thd*1.0, thd*0.4]
        self.bar_espectro.setOpts(height=h_vals)
        self.plt_espectro.getViewBox().setLimits(xMin=0, xMax=9, yMin=0, yMax=110)
        if not self.zoomed_esp:
            self.plt_espectro.setXRange(0, 9, padding=0.05); self.plt_espectro.setYRange(0, 110, padding=0)

        x_st = np.arange(len(self.motor.hist_v))
        self.plt_st_v[1].setData(x_st, self.motor.hist_v)
        self.plt_st_i[1].setData(x_st, self.motor.hist_i)
        self.plt_st_vib[1].setData(x_st, self.motor.hist_vib)
        for pw, data, name in [(self.plt_st_v[0], self.motor.hist_v, 'st_v'),
                                (self.plt_st_i[0], self.motor.hist_i, 'st_i'),
                                (self.plt_st_vib[0], self.motor.hist_vib, 'st_vib')]:
            setattr(self, f"auto_{name}", True); mn2=np.mean(data); r2=max(mn2*0.1, 0.1)
            pw.getViewBox().setLimits(xMin=0, xMax=len(x_st), yMin=mn2-r2, yMax=mn2+r2)
            if not getattr(self, f"zoomed_{name}"): pw.setXRange(0, len(x_st), padding=0); pw.setYRange(mn2-r2, mn2+r2, padding=0)
            setattr(self, f"auto_{name}", False)

        scale = 150.0/2500.0; ps=p*scale; qs=q*scale
        self.ln_p.setData([0, ps], [0, 0]); self.ln_q.setData([ps, ps], [0, qs]); self.ln_s.setData([0, ps], [0, qs])
        self.txt_P.setPos(ps/2, 5); self.txt_Q.setPos(ps+5, qs/2); self.txt_S.setPos(ps/2-15, qs/2+10)
        self.auto_tri = True
        xmax_tri=max(200.0, ps+20); ymax_tri=max(150.0, qs+20)
        self.plt_tri.getViewBox().setLimits(xMin=-20, xMax=xmax_tri, yMin=-20, yMax=ymax_tri)
        if not self.zoomed_tri:
            self.plt_tri.setXRange(-20, xmax_tri, padding=0); self.plt_tri.setYRange(-20, ymax_tri, padding=0)
        self.auto_tri = False

        x_aud = self.motor.fast_ts; tiempo_actual = x_aud[-1]; x_aud_rel = x_aud - tiempo_actual
        audit_salud_visual = self.motor.fast_salud*100 - 80
        audit_vib_visual   = self.motor.fast_vib*10 - 10
        audit_thd_visual   = self.motor.fast_thd*2
        self.ln_aud_salud.setData(x_aud_rel, audit_salud_visual)
        self.ln_aud_thd.setData(x_aud_rel, audit_thd_visual)
        self.ln_aud_vib.setData(x_aud_rel, audit_vib_visual)
        t_min=-60.0; t_max=0.0
        s_min=np.min([np.min(audit_salud_visual), np.min(audit_thd_visual), np.min(audit_vib_visual)])
        s_max=np.max([np.max(audit_salud_visual), np.max(audit_thd_visual), np.max(audit_vib_visual)])
        s_rng=max(5.0, (s_max-s_min)*0.05); y_min_lim=s_min-s_rng; y_max_lim=s_max+s_rng
        self.auto_audit = True
        self.plt_audit.getViewBox().setLimits(xMin=t_min, xMax=t_max, yMin=y_min_lim, yMax=y_max_lim)
        if not self.zoomed_audit:
            self.plt_audit.setXRange(t_min, t_max, padding=0); self.plt_audit.setYRange(y_min_lim, y_max_lim, padding=0)
        self.auto_audit = False

    # BUCLE LENTO (1s)
    def _loop_slow(self):
        v=self.motor.current_v; i=self.motor.current_i; p=self.motor.current_p
        q=self.motor.current_q; s=self.motor.current_s; pf=self.motor.current_pf
        thd=self.motor.current_thd; vib=self.motor.current_vib
        freq=self.motor.current_freq; temp=self.motor.current_temp
        salud=self.motor.current_salud; ts=datetime.now().strftime("%H:%M:%S")

        # INFERENCIA TENSORFLOW
        import threading
        features = [v, i, p, q, pf, thd, vib, temp]

        if not self.tf_brain.is_trained and \
           len(self.tf_brain.training_buf) >= self.tf_brain.MIN_SAMPLES and \
           not self._tf_entrenando:
            self._tf_entrenando = True
            self.tf_brain.training_buf.append(features)
            self.tf_brain.samples_seen += 1
            def entrenar_en_hilo():
                self.tf_brain._train_initial()
                self.tf_brain.is_trained = True
                self._tf_entrenando = False
            threading.Thread(target=entrenar_en_hilo, daemon=True).start()
            self.tf_result = self.tf_brain._make_result("⚙️ Entrenando modelos TF en segundo plano...")
        elif self._tf_entrenando:
            self.tf_result = self.tf_brain._make_result("⚙️ Entrenamiento en curso, un momento...")
        else:
            self.tf_result = self.tf_brain.update(features)
        tf_r = self.tf_result

        # --- NUEVA LÓGICA: Traducir datos a nombres de fallas ---
        anomalias_detectadas = []
        
        # 1. Chequeo de límites físicos
        if v < 110 or v > 140: anomalias_detectadas.append("Tensión fuera de rango nominal")
        if i > 50:             anomalias_detectadas.append("Sobrecorriente detectada")
        if vib > 4:            anomalias_detectadas.append("Vibración mecánica fuera de norma")
        if thd > 8:            anomalias_detectadas.append("Distorsión armónica elevada (THD)")
        if temp > 65:          anomalias_detectadas.append("Sobrecalentamiento en estator")
        if pf < 0.85 and freq > 0: anomalias_detectadas.append("Factor de potencia ineficiente")
        
        # 2. Chequeo de lógica de IA (TensorFlow)
        if tf_r['ae_loss'] > self.tf_brain.threshold_ae:
            anomalias_detectadas.append("Firma eléctrica anómala (Autoencoder)")
        if tf_r['pred_salud'] < 75:
            anomalias_detectadas.append("Degradación de salud proyectada (LSTM)")
        
        # Unimos la lista en un solo texto con viñetas
        txt_anom_list = "\n  • ".join(anomalias_detectadas) if anomalias_detectadas else "Ninguna (Sistema Nominal)"

        # Actualizar indicador en topbar
        estado_color = {
            "APRENDIENDO": C_ACCENT, "NOMINAL": C_GOOD,
            "PREVENCIÓN": C_ORANGE, "ANOMALÍA AE": C_CRIT,
            "ANOMALÍA IF": C_CRIT, "ANOMALÍA CRÍTICA": C_CRIT,
        }.get(tf_r['estado'], C_TEXT_M)
        self.lbl_tf_estado.setText(f"◈ MEC: {tf_r['estado']} | Confianza: {tf_r['confianza']:.0f}%")
        self.lbl_tf_estado.setStyleSheet(f"color:{estado_color}; font-size:10px; font-weight:bold;")

        # Actualizar sidebar TF
        iso_sc_side = self.tf_brain.iso_forest.decision_function(
            [[self.motor.current_v, self.motor.current_i, self.motor.current_p,
              self.motor.current_q, self.motor.current_pf, self.motor.current_thd,
              self.motor.current_vib, self.motor.current_temp]]
        )[0] if self.tf_brain.is_trained else 0.0

        ae_val   = tf_r['hist_ae'][-1]   if tf_r['hist_ae']   else tf_r['ae_loss']
        sal_val  = tf_r['hist_sal'][-1]  if tf_r['hist_sal']  else tf_r['pred_salud']
        anom_val = tf_r['hist_anom'][-1] if tf_r['hist_anom'] else tf_r['anomaly_score']
        nivel_anom = "❌ CRÍTICO" if anom_val > 3 else ("⚠️ AVISO" if anom_val > 1.5 else "✅ NORMAL")
        nivel_ae = "❌" if (ae_val > self.tf_brain.threshold_ae) else ("⚠️" if ae_val > self.tf_brain.threshold_ae * 0.75 else "✅")
        nivel_sal = "❌" if sal_val < 40 else ("⚠️" if sal_val < 75 else "✅")

        texto_cuadro = (
            f"◤ MEC · NÚCLEO\n"
            f"{'─'*24}\n"
            f"Estado   : {tf_r['estado']}\n"
            f"Muestras : {tf_r['samples']}\n"
            f"{'─'*24}\n"
            f"AE Loss  : {nivel_ae} {ae_val:.5f}\n"
            f"Umbral   :    {self.tf_brain.threshold_ae:.5f}\n"
            f"LSTM     : {nivel_sal} {sal_val:.1f}%\n"
            f"IF Score :    {iso_sc_side:.3f}\n"
            f"Anomalía : {nivel_anom}\n"
            f"σ        :    {anom_val:.3f}\n"
            f"{'─'*24}\n"
            f"Anomalías: {tf_r['n_anomalias']} detectadas"
        )
        self.lbl_tf_info.setText(texto_cuadro)

        # VERIFICACIÓN MATEMÁTICA
        ok_list, advertencias, errores, resumen_mat = self.tf_brain.get_verificacion_matematica(
            v, i, p, q, s, pf, thd, vib, freq, temp, salud
        )

        # LLENADO DE TABLAS PRINCIPALES
        self.motor.tick_audit(v, i, p, q, s, pf, vib, freq, thd, temp, salud)
        n = len(self.motor.audit_labels); start = max(0, n-7); recent = list(range(start, n))[::-1]
        d_rms, d_osc, d_tri, d_aud = [], [], [], []
        for m in recent:
            t_l=self.motor.audit_labels[m]; v_v=self.motor.audit_v[m]; i_v=self.motor.audit_i[m]
            f_g=self.motor.current_f_rms_global
            d_rms.append([t_l, f"{f_g:.2f}", f"{v_v:.1f}", f"{i_v:.2f}",
                          f"{self.motor.audit_p[m]:.0f}", f"{self.motor.audit_q[m]:.0f}",
                          f"{self.motor.audit_s[m]:.0f}", f"{self.motor.audit_pf[m]:.2f}",
                          f"{self.motor.audit_thd[m]:.1f}", f"{self.motor.audit_vib[m]:.2f}",
                          f"{self.motor.audit_freq[m]:.1f}", f"{self.motor.audit_temp[m]:.1f}"])
            d_tri.append([t_l, f"{v_v:.1f}", f"{self.motor.audit_salud[m]*100:.1f}%", f"{thd:.1f}%"])
            d_osc.append([t_l, f"{v_v:.1f}", f"{i_v:.2f}", f"{p:.0f}", f"{freq:.1f}"])
            d_aud.append([t_l, f"{v_v:.1f}", f"{i_v:.2f}", f"{p:.0f}", f"{vib:.2f}", f"{freq:.1f}", f"{thd:.1f}", f"{temp:.1f}", f"{self.motor.audit_salud[m]*100:.1f}%"])
        self._fill_table(self.tbl_rms, d_rms); self._fill_table(self.tbl_tri, d_tri)
        self._fill_table(self.tbl_osc, d_osc); self._fill_table(self.tbl_audit, d_aud)

        d_esp = []
        if n > 0:
            ult_esp = self.motor.audit_espectro[-1]; fund = ult_esp[0]
            for idx, mag in enumerate(ult_esp):
                cont = (mag/fund*10) if fund != 0 else 0
                d_esp.append([str([1,2,3,4,5,6,7,8][idx]), f"{mag:.1f}%", f"{cont:.2f}"])
        self._fill_table(self.tbl_espectro, d_esp)

        # LOG DEL SISTEMA EXPERTO (TF)
        torque   = (p / (2*np.pi*freq)) if freq > 0 else 0
        mtbf_est = max(0, 8000*(salud**2) - (temp*15))
        tipo_corriente = "ALTERNA (AC)" if freq > 0 else "DIRECTA (DC)"

        riesgo_thd  = "CRÍTICO" if thd>15 else ("ELEVADO" if thd>8 else ("NORMAL" if thd>3 else "ÓPTIMO"))
        riesgo_vib  = "CRÍTICO" if vib>8  else ("ELEVADO" if vib>4  else ("NORMAL" if vib>2  else "ÓPTIMO"))
        riesgo_temp = "CRÍTICO" if temp>80 else ("ELEVADO" if temp>65 else ("NORMAL" if temp>50 else "ÓPTIMO"))
        riesgo_pf   = "CRÍTICO" if pf<0.7  else ("BAJO" if pf<0.85   else ("ACEPTABLE" if pf<0.92 else "ÓPTIMO"))
        riesgo_sal  = "CRÍTICA" if salud<0.5 else ("BAJA" if salud<0.75 else ("ACEPTABLE" if salud<0.9 else "EXCELENTE"))

        # Sección: Verificación matemática
        def fmt_lista(lst): return "\n".join(f"  {x}" for x in lst) if lst else "  (ninguno)"

        # Predicciones
        pred_txt = []
        if thd > 8: pred_txt.append(f"THD={thd:.1f}%: Envejecimiento acelerado del aislamiento.")
        if vib > 4: pred_txt.append(f"Vib={vib:.2f}mm/s: Fatiga de rodamientos estimada.")
        if temp > 65: pred_txt.append(f"Temp={temp:.1f}°C: Degradación del barniz aislante.")
        if pf < 0.75 and freq > 0: pred_txt.append(f"PF={pf:.2f}: Penalización económica CFE.")
        if not pred_txt: pred_txt.append(f"MTBF estimado: {mtbf_est:.0f}h de operación confiable.")

        recomend = []
        if thd > 8: recomend.append("Instalar filtro de armónicos activo o pasivo.")
        if vib > 4: recomend.append("Balanceo dinámico del rotor. Revisar alineación.")
        if temp > 65: recomend.append("Mejorar ventilación. Reducir carga.")
        if pf < 0.85 and freq > 0:
            C_cap = (q/(2*np.pi*freq*v**2))*1e6 if freq > 0 else 0
            recomend.append(f"Agregar capacitor C≈{C_cap:.1f}µF para corregir PF.")
        if not recomend: recomend.append("Sin acciones correctivas requeridas.")

        def ind_v(val):
            desv = abs(val - 127.0) / 127.0 * 100
            if desv > 15: return f"❌ {val:.2f}V — FUERA DE RANGO CRÍTICO"
            elif desv > 8: return f"⚠️  {val:.2f}V — Desviación elevada ({desv:.1f}%)"
            else: return f"   {val:.2f}V"

        def ind_i(val):
            if val > 80: return f"❌ {val:.2f}A — SOBRECORRIENTE CRÍTICA"
            elif val > 50: return f"⚠️  {val:.2f}A — Corriente elevada"
            else: return f"   {val:.2f}A"

        def ind_p(val):
            if val > 12000: return f"⚠️  {val:.0f}W — Potencia elevada"
            else: return f"   {val:.0f}W"

        def ind_pf(val):
            if val < 0.7: return f"❌ {val:.3f} — FACTOR DE POTENCIA CRÍTICO"
            elif val < 0.85: return f"⚠️  {val:.3f} — PF bajo, penalización CFE posible"
            elif val < 0.92: return f"⚠️  {val:.3f} — Aceptable, pero mejorable"
            else: return f"   {val:.3f}"

        def ind_thd(val):
            if val > 15: return f"❌ {val:.1f}% — VIOLA IEEE 519"
            elif val > 8: return f"⚠️  {val:.1f}% — THD elevado"
            elif val > 5: return f"⚠️  {val:.1f}% — Sobre recomendado"
            else: return f"   {val:.1f}%"

        def ind_vib(val):
            if val > 8: return f"❌ {val:.2f}mm/s — PELIGROSO ISO 10816 Zona C/D"
            elif val > 4: return f"⚠️  {val:.2f}mm/s — Zona B, monitorear"
            else: return f"   {val:.2f}mm/s"

        def ind_temp(val):
            if val > 80: return f"❌ {val:.1f}°C — TEMPERATURA CRÍTICA"
            elif val > 65: return f"⚠️  {val:.1f}°C — Temperatura elevada"
            elif val > 50: return f"⚠️  {val:.1f}°C — Zona de monitoreo"
            else: return f"   {val:.1f}°C"

        def ind_freq(val):
            if val <= 0: return f"❌ {val:.1f}Hz — SIN FRECUENCIA"
            elif val < 55 or val > 65: return f"❌ {val:.1f}Hz — FUERA DE RANGO"
            else: return f"   {val:.1f}Hz"

        def ind_salud(val):
            pct = val * 100
            if pct < 50: return f"❌ {pct:.1f}% — SALUD CRÍTICA"
            elif pct < 75: return f"⚠️  {pct:.1f}% — Salud baja"
            elif pct < 90: return f"⚠️  {pct:.1f}% — Aceptable"
            else: return f"   {pct:.1f}%"

        def ind_lstm(val):
            if val < 40: return f"❌ {val:.1f}% — LSTM: deterioro severo"
            elif val < 65: return f"⚠️  {val:.1f}% — LSTM: tendencia negativa"
            elif val < 85: return f"⚠️  {val:.1f}% — LSTM: salud moderada"
            else: return f"   {val:.1f}%"

        def ind_ae(val, umbral):
            ratio = val / max(umbral, 1e-9)
            if ratio > 1.0: return f"❌ {val:.5f} — SUPERA UMBRAL (×{ratio:.2f})"
            elif ratio > 0.75: return f"⚠️  {val:.5f} — Acercándose ({ratio*100:.0f}%)"
            else: return f"   {val:.5f}"

        def ind_anom(val):
            if val > 3: return f"❌ {val:.3f}σ — ANOMALÍA CONFIRMADA"
            elif val > 1.5: return f"⚠️  {val:.3f}σ — Desviación significativa"
            else: return f"   {val:.3f}σ"

        # Análisis narrativo TF en tiempo real
        if not self.tf_brain.is_trained:
            analisis_tf = (
                f"  Ingeniero, aún estoy en fase de aprendizaje.\n"
                f"  Llevo {tf_r['samples']} de {self.tf_brain.MIN_SAMPLES} muestras necesarias.\n"
                f"  Por favor mantenga el sistema en condiciones\n"
                f"  nominales durante esta etapa. No puedo hacer\n"
                f"  inferencias confiables todavía."
            )
        else:
            ae_loss  = tf_r['ae_loss']
            pred_sal = tf_r['pred_salud']
            score_s  = tf_r['anomaly_score']
            umbral   = self.tf_brain.threshold_ae
            estado_tf = tf_r['estado']

            # LO QUE VEO
            if ae_loss < umbral * 0.5:
                veo = (f"La firma eléctrica es muy consistente con\n"
                       f"  lo que aprendí como normal. AE-Loss={ae_loss:.5f},\n"
                       f"  solo {ae_loss/umbral*100:.0f}% del umbral calibrado.")
            elif ae_loss < umbral:
                veo = (f"La señal está dentro del rango normal pero\n"
                       f"  con más variabilidad de lo habitual.\n"
                       f"  AE-Loss={ae_loss:.5f} ({ae_loss/umbral*100:.0f}% del umbral).")
            else:
                veo = (f"Detecto una desviación real en la firma\n"
                       f"  eléctrica. El error de reconstrucción\n"
                       f"  ({ae_loss:.5f}) supera mi umbral ({umbral:.5f}).")

            # LO QUE NO VEO
            if pred_sal > 85:
                no_veo = (f"No detecto señales ocultas de falla inminente.\n"
                          f"  El LSTM proyecta {pred_sal:.1f}% de salud,\n"
                          f"  lo que indica continuidad operativa estable.")
            elif pred_sal > 60:
                no_veo = (f"El LSTM proyecta {pred_sal:.1f}% de salud. Hay\n"
                          f"  algo en la tendencia histórica que no termina\n"
                          f"  de cuadrar. Puede ser ruido o una deriva.")
            else:
                no_veo = (f"⚠️ El LSTM proyecta solo {pred_sal:.1f}% de salud.\n"
                          f"  Hay patrones en el historial que me preocupan\n"
                          f"  y que no son visibles en valores instantáneos.")

            # LO QUE ESTÁ PASANDO
            if estado_tf == "NOMINAL":
                pasando = (f"Los tres modelos coinciden: el punto de\n"
                           f"  operación actual es normal. Autoencoder,\n"
                           f"  LSTM e IsolationForest están de acuerdo.")
            elif estado_tf == "PREVENCIÓN":
                pasando = (f"Hay una desviación leve (σ={score_s:.2f}) que\n"
                           f"  no alcanza umbral de anomalía, pero se aleja\n"
                           f"  del centro de la distribución aprendida.")
            elif "AE" in estado_tf:
                pasando = (f"El Autoencoder no puede reconstruir bien\n"
                           f"  la señal actual. El motor opera fuera del\n"
                           f"  espacio de estados que aprendí como normal.")
            elif "IF" in estado_tf:
                pasando = (f"IsolationForest clasifica este punto como\n"
                           f"  outlier. Es una señal temprana que no\n"
                           f"  debe ignorarse aunque el AE aún no dispare.")
            elif "CRÍTICA" in estado_tf:
                pasando = (f"🚨 ALERTA: Dos detectores coinciden.\n"
                           f"  Hay una anomalía real en curso. No es\n"
                           f"  ruido — es una desviación sostenida.")
            else:
                pasando = f"  Estado: {estado_tf}. Continúo monitoreando."

            # LO QUE PODRÍA PASAR
            riesgos = []
            if thd > 8:   riesgos.append(f"THD={thd:.1f}%: envejecimiento del aislamiento.")
            if vib > 4:   riesgos.append(f"Vib={vib:.2f}mm/s: fatiga de rodamientos.")
            if temp > 65: riesgos.append(f"Temp={temp:.1f}°C: degradación del barniz.")
            if pf < 0.85 and freq > 0: riesgos.append(f"PF={pf:.3f}: cargos por energía reactiva.")
            if pred_sal < 70: riesgos.append(f"LSTM {pred_sal:.1f}%: posible falla funcional próxima.")
            if not riesgos: riesgos.append(f"Sin riesgos identificados. MTBF≈{mtbf_est:.0f}h.")

            # OPINIÓN PERSONAL
            n_adv = len(advertencias)
            n_err = len(errores)
            if estado_tf == "NOMINAL" and n_adv == 0 and n_err == 0:
                opinion = (f"Ingeniero, el motor está en excelentes\n"
                           f"  condiciones. No requiere acción inmediata.")
            elif estado_tf == "NOMINAL" and n_adv > 0:
                opinion = (f"El sistema opera bien, pero {n_adv} parámetro(s)\n"
                           f"  merecen atención en la próxima ventana\n"
                           f"  de mantenimiento preventivo.")
            elif "PREVENCIÓN" in estado_tf:
                opinion = (f"Le recomiendo mantener vigilancia activa.\n"
                           f"  No es urgente, pero no debe ignorarse\n"
                           f"  por más de 2 o 3 turnos de operación.")
            else:
                opinion = (f"Mi recomendación es actuar pronto.\n"
                           f"  Cuando tres algoritmos independientes\n"
                           f"  coinciden en una desviación, es real.")

            analisis_tf = (
                f"  LO QUE VEO:\n  {veo}\n\n"
                f"  LO QUE NO VEO:\n  {no_veo}\n\n"
                f"  LO QUE ESTÁ PASANDO:\n  {pasando}\n\n"
                f"  LO QUE PODRÍA PASAR:\n"
                + "\n".join(f"  • {r}" for r in riesgos) +
                f"\n\n  MI OPINIÓN:\n  {opinion}"
            )

        # Sección 1: Estado + Métricas unificados
        sec_estado_metricas = (
            f"Hora: {ts} | MEC [{tipo_corriente}]\n"
            f"ESTADO: {tf_r['estado']} | Confianza: {tf_r['confianza']:.0f}% | Muestras: {tf_r['samples']}\n"
            f"─────────────────────────────────────────\n"
            f"ANOMALÍAS ACTIVAS:\n"
            f"  • {txt_anom_list}\n"
            f"─────────────────────────────────────────\n"
            f"Tensión     : {ind_v(v)}\n"
            f"Corriente   : {ind_i(i)}\n"
            f"Potencia    : {ind_p(p)}\n"
            f"Factor Pot. : {ind_pf(pf)}\n"
            f"THD         : {ind_thd(thd)}\n"
            f"Vibración   : {ind_vib(vib)}\n"
            f"Temperatura : {ind_temp(temp)}\n"
            f"Frecuencia  : {ind_freq(freq)}\n"
            f"Salud Real  : {ind_salud(salud)}\n"
            f"Salud LSTM  : {ind_lstm(tf_r['pred_salud'])}\n"
            f"V_RMS       :    {self.motor.current_v_rms:.3f}V\n"
            f"I_RMS       :    {self.motor.current_i_rms:.3f}A\n"
            f"Torque      :    {torque:.2f}N·m\n"
            f"MTBF est.   :    {mtbf_est:.0f}h"
        )
        vbar1 = self.txt_sec_estado.verticalScrollBar(); pos1 = vbar1.value()
        self.txt_sec_estado.setHtml(self._mec_html("MEC // TELEMETRÍA EN VIVO", sec_estado_metricas, C_HUD))
        vbar1.setValue(pos1)

        # Sección 2: Análisis narrativo — lo escribe MEC (en primera persona)
        cuerpo_analisis = (
            f"{analisis_tf}\n\n"
            f"✅ RECOMENDACIONES: {' | '.join(recomend)}"
        )
        vbar2 = self.txt_sec_analisis.verticalScrollBar(); pos2 = vbar2.value()
        self.txt_sec_analisis.setHtml(self._mec_html("MEC // ANÁLISIS EN TIEMPO REAL", cuerpo_analisis, C_GOOD))
        vbar2.setValue(pos2)

        # Texto plano para el reporte de Excel (conserva el contenido completo)
        log_completo_plano = "MEC — ANÁLISIS EN TIEMPO REAL\n" + cuerpo_analisis
        self._texto_excel = sec_estado_metricas + "\n\n" + log_completo_plano

    # EXPORTAR EXCEL
    def _exportar_excel(self):
        m = self.motor; n = len(m.audit_labels)
        if n == 0: self.lbl_excel_status.setText("⚠ Sin datos aún."); return
        MAX_RECORDS = 50; start = max(0, n-MAX_RECORDS); idx_list = list(range(start, n))
        ts_file = datetime.now().strftime("%d-%m-%y-%I.%M %p")
        filename = f"Reporte MEC {ts_file}.xlsx"
        EXPORT_DIR = r"C:\Users\Amado\OneDrive\Desktop\ProyectoMEC\Reportes"
        os.makedirs(EXPORT_DIR, exist_ok=True); path = os.path.join(EXPORT_DIR, filename)
        wb = Workbook(); ws = wb.active; ws.title = "Auditoría Técnica"; ws.sheet_view.showGridLines = True
        title_font = Font(name="Segoe UI", bold=True, color="000000", size=14)
        meta_font  = Font(name="Consolas", bold=False, color="000000", size=9)
        hdr_font   = Font(name="Consolas", bold=True, color="000000", size=7.5)
        reg_font   = Font(name="Consolas", bold=False, color="000000", size=7.5)
        ia_title_f = Font(name="Consolas", bold=True, color="000000", size=10)
        ia_text_f  = Font(name="Consolas", bold=False, color="000000", size=9)
        center_align = Alignment(horizontal="center", vertical="center")
        left_top_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        header_fill = PatternFill("solid", fgColor="00B050")
        ia_fill     = PatternFill("solid", fgColor="EAF2F8")
        tf_fill     = PatternFill("solid", fgColor="EDE7F6")

        ws.merge_cells("A1:Z1")
        ws["A1"].value = "MEC INDUSTRIAL ANALYTICS PRO — REPORTE TÉCNICO (TensorFlow AI Engine)"
        ws["A1"].font = title_font; ws["A1"].alignment = center_align; ws["A1"].fill = header_fill
        ws.merge_cells("A2:Z2")
        fecha_h = datetime.now().strftime('%d/%m/%Y %I:%M:%S %p')
        ws["A2"].value = f"EQUIPO: MEC-01 | GENERADO: {fecha_h} | REGISTROS: {len(idx_list)} | TF-Estado: {self.tf_result['estado']} | Anomalías: {self.tf_result['n_anomalias']}"
        ws["A2"].font = meta_font; ws["A2"].alignment = center_align; ws["A2"].fill = header_fill

        HEADERS = ["Timestamp","Tensión(V)","Corr(A)","P.Act(W)","P.Rea(VAr)","P.Apa(VA)","F.P.","THD(%)","Vib(mm/s)","Freq(Hz)","Temp(°C)","f_RMS","Salud(%)","H1(%)","H3(%)","H5(%)","H7(%)","H9(%)"]
        for col_idx, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = hdr_font; cell.border = thin_border; cell.alignment = center_align; cell.fill = header_fill

        row_num = 5
        for mem_idx in idx_list:
            lbl_=m.audit_labels[mem_idx]; v_=round(m.audit_v[mem_idx],2); i_=round(m.audit_i[mem_idx],3)
            pf_=round(m.audit_pf[mem_idx],4); thd_=round(m.audit_thd[mem_idx],2)
            vib_=round(m.audit_vib[mem_idx],3); freq_=round(m.audit_freq[mem_idx],1); temp_=round(m.audit_temp[mem_idx],1)
            r = row_num
            row_data = [lbl_, v_, i_, f"=B{r}*C{r}*G{r}", f"=SQRT(MAX(0,F{r}^2-D{r}^2))", f"=B{r}*C{r}",
                        pf_, thd_, vib_, freq_, temp_,
                        f"=SQRT((B{r}^2+C{r}^2+D{r}^2+E{r}^2+F{r}^2+G{r}^2+H{r}^2+I{r}^2+J{r}^2+K{r}^2)/10)",
                        f"=MAX(0,MIN(1,1-(I{r}/15)*0.7-(H{r}/30)*0.3))*100",
                        "=100", f"=H{r}*2.0", f"=H{r}*1.5", f"=H{r}*1.0", f"=H{r}*0.7"]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = reg_font; cell.border = thin_border; cell.alignment = center_align
                if isinstance(val, str) and val.startswith("="): cell.number_format = '0.00'
            row_num += 1

        row_num += 1
        cell_avg = ws.cell(row=row_num, column=1, value="PROMEDIO")
        cell_avg.font = hdr_font; cell_avg.border = thin_border; cell_avg.alignment = center_align; cell_avg.fill = header_fill
        for col_idx in range(2, 19):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = f"=AVERAGE({col_letter}5:{col_letter}{row_num-2})"
            cell.font = hdr_font; cell.border = thin_border; cell.alignment = center_align; cell.number_format = '0.00'; cell.fill = header_fill

        # Sección IA / TF
        col_ia_start = 20; col_ia_end = 23
        ws.merge_cells(start_row=4, start_column=col_ia_start, end_row=4, end_column=col_ia_end)
        header_ia = ws.cell(row=4, column=col_ia_start, value="ANÁLISIS DEL SISTEMA EXPERTO — TENSORFLOW AI")
        header_ia.font = ia_title_f; header_ia.alignment = center_align; header_ia.fill = tf_fill
        ia_end_row = max(row_num, 56)
        ws.merge_cells(start_row=5, start_column=col_ia_start, end_row=ia_end_row, end_column=col_ia_end)
        reporte_box = ws.cell(row=5, column=col_ia_start)
        reporte_box.value = getattr(self, '_texto_excel', self.txt_reporte.toPlainText())
        for r in range(4, ia_end_row+1):
            for c in range(col_ia_start, col_ia_end+1):
                cell = ws.cell(row=r, column=c); cell.border = thin_border; cell.fill = tf_fill

        ws.column_dimensions['S'].width = 3; ws.column_dimensions['T'].width = 3
        for i in range(col_ia_start, col_ia_end+1): ws.column_dimensions[get_column_letter(i)].width = 15

        try:
            wb.save(path)
            self.lbl_excel_status.setText("✅ Reporte TF Finalizado")
            QTimer.singleShot(3000, lambda: self.lbl_excel_status.setText(""))
        except Exception as ex:
            self.lbl_excel_status.setText(f"❌ Error: {str(ex)}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    app.setFont(QFont('Segoe UI'))
    w = ConsolaCFE()
    w.show()
    sys.exit(app.exec())